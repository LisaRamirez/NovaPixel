"""Acciones de admin reutilizables para exportar querysets a Excel o PDF.

Cada ModelAdmin que quiera exportar hereda de ExportMixin y define
`export_fields`: una lista de tuplas (encabezado, atributo_o_callable). Un
string se resuelve como atributo/método del objeto (soporta "get_x_display"
y rutas con puntos tipo "product.name"); un callable recibe el objeto y
devuelve el valor a mostrar.
"""

import io

from django.contrib import admin
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Table, TableStyle

NOVA_PURPLE = "9B5DE5"


def _resolve(obj, field):
    if callable(field):
        return field(obj)
    value = obj
    for part in field.split("."):
        value = getattr(value, part, "")
        if callable(value):
            value = value()
    return "" if value is None else str(value)


class ExportMixin:
    """Agrega las acciones "Exportar a Excel" y "Exportar a PDF" a un
    ModelAdmin. Requiere `export_fields = [(header, field), ...]` en la
    subclase; `export_title` es opcional (por defecto usa el verbose_name)."""

    export_fields = []
    export_title = None
    actions = ["export_as_excel", "export_as_pdf"]

    def get_export_rows(self, queryset):
        headers = [header for header, _ in self.export_fields]
        rows = [[_resolve(obj, field) for _, field in self.export_fields] for obj in queryset]
        return headers, rows

    @admin.action(description="Exportar seleccionados a Excel (.xlsx)")
    def export_as_excel(self, request, queryset):
        headers, rows = self.get_export_rows(queryset)

        wb = Workbook()
        ws = wb.active
        ws.title = str(self.export_title or self.model._meta.verbose_name_plural)[:31]

        header_fill = PatternFill(start_color=NOVA_PURPLE, end_color=NOVA_PURPLE, fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        ws.append(headers)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        for row in rows:
            ws.append(row)

        for col in ws.columns:
            width = max((len(str(c.value)) if c.value is not None else 0) for c in col) + 2
            ws.column_dimensions[col[0].column_letter].width = min(width, 40)

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{self.model._meta.model_name}.xlsx"'
        wb.save(response)
        return response

    @admin.action(description="Exportar seleccionados a PDF")
    def export_as_pdf(self, request, queryset):
        headers, rows = self.get_export_rows(queryset)

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), topMargin=1.5 * cm, bottomMargin=1.5 * cm)
        styles = getSampleStyleSheet()
        title = str(self.export_title or self.model._meta.verbose_name_plural).capitalize()

        table = Table([headers] + rows, repeatRows=1)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(f"#{NOVA_PURPLE}")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cccccc")),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f5f5f5")]),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ]
            )
        )
        doc.build([Paragraph(f"NovaPixel — {title}", styles["Title"]), table])

        response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
        response["Content-Disposition"] = f'attachment; filename="{self.model._meta.model_name}.pdf"'
        buffer.close()
        return response
